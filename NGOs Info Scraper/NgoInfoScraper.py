from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

pageNo = 5


def getSheet():
    sheet_name = f"Ngo{pageNo}"
    if sheet_name in workbook.sheetnames:
        raise ValueError("The sheet is already present. Kindly delete the sheet first!")
    return workbook.create_sheet(title=sheet_name)


workbook = load_workbook("NgosData.xlsx")
new_sheet = getSheet()

driver = webdriver.Chrome()
driver.get(
    f"https://ngodarpan.gov.in/index.php/home/statewise_ngo/5466/19/{pageNo}?per_page=100"
)

name = []

uniqueId = []
regNo = []
typeOfNgo = []

head = []
designation = []

address = []
city = []
state = []
mobile = []
webUrl = []
email = []


def webElement(path):
    return driver.find_element(
        by=By.XPATH,
        value=path,
    )


def webScrap(path):
    return webElement(path).text


for i in range(1, 101):

    nameElement = webElement(
        f"/html/body/div[9]/div[1]/div[3]/div/div/div[2]/table/tbody/tr[{i}]/td[2]/a"
    )

    nameElement.click()
    name.append(nameElement.text)

    time.sleep(4)

    uniqueId.append(webScrap('//*[@id="UniqueID"]'))
    regNo.append(webScrap('//*[@id="ngo_regno"]'))
    typeOfNgo.append(webScrap('//*[@id="ngo_type"]'))

    head.append(webScrap('//*[@id="member_table"]/tbody/tr[2]/td[1]'))
    designation.append(webScrap('//*[@id="member_table"]/tbody/tr[2]/td[2]'))

    mobile.append(webScrap('//*[@id="mobile_n"]'))
    email.append(webScrap('//*[@id="email_n"]'))
    webUrl.append(webScrap('//*[@id="ngo_web_url"]'))

    address.append(webScrap('//*[@id="address"]'))
    city.append(webScrap('//*[@id="city"]'))
    state.append(webScrap('//*[@id="state_p_ngo"]'))

    webElement('//*[@id="ngo_info_modal"]/div[2]/div/div[1]/button/span').click()

    time.sleep(3)

    print(i)


driver.close()
driver.quit()

print(name)


def getEmail(email):
    if email == "Not Available":
        return "null"
    email = email.replace("(at)", "@")
    email = email.replace("[dot]", ".")
    return email.lower()


def getUrl(url):
    if url == "Not Available":
        return "null"
    return url


headers = [
    "no.",
    "name",
    "uniqueID",
    "regNo",
    "type",
    "head",
    "designation",
    "mobile",
    "email",
    "webUrl",
    "registered",
    "address",
    "city",
    "state",
    "fullAddress",
]

for col, header in enumerate(headers, 1):
    new_sheet.cell(row=1, column=col, value=header)

for i in range(100):
    new_sheet.cell(row=i + 2, column=1, value=i + 1)
    new_sheet.cell(row=i + 2, column=2, value=name[i].title())
    new_sheet.cell(row=i + 2, column=3, value=uniqueId[i])
    new_sheet.cell(row=i + 2, column=4, value=regNo[i])
    new_sheet.cell(row=i + 2, column=5, value=typeOfNgo[i])
    new_sheet.cell(row=i + 2, column=6, value=head[i].title())
    new_sheet.cell(row=i + 2, column=7, value=designation[i])
    new_sheet.cell(row=i + 2, column=8, value=f"+91{mobile[i]}")
    new_sheet.cell(row=i + 2, column=9, value=getEmail(email[i]))
    new_sheet.cell(row=i + 2, column=10, value=getUrl(webUrl[i]))
    new_sheet.cell(row=i + 2, column=11, value=False)
    new_sheet.cell(row=i + 2, column=12, value=address[i].title())
    new_sheet.cell(row=i + 2, column=13, value=city[i].title())
    new_sheet.cell(row=i + 2, column=14, value=state[i].title())
    new_sheet.cell(row=i + 2, column=15, value=f"{address[i]}, {city[i]}, {state[i]}")

workbook.save("NgosData.xlsx")
